import openpyxl
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout

FILE_NAME = "data_nilai.xlsx"

class NilaiTahfizhApp(App):
    def simpan_nilai(self):
        nama = self.nama_spinner.text
        sabaq = self.sabaq_input.text
        sabqi = self.sabqi_input.text
        tahsin = self.tahsin_input.text
        suluk = self.suluk_input.text

        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active

        # Mencari baris siswa
        siswa_baris = None
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == nama:
                siswa_baris = row
                break

        if siswa_baris is None:
            siswa_baris = ws.max_row + 1
            ws.cell(row=siswa_baris, column=1, value=nama)

        # Simpan nilai di kolom yang sesuai
        tanggal_kolom = 2  # Sesuaikan dengan posisi tanggal
        ws.cell(row=siswa_baris, column=tanggal_kolom, value=sabaq)
        ws.cell(row=siswa_baris, column=tanggal_kolom+1, value=sabqi)
        ws.cell(row=siswa_baris, column=tanggal_kolom+2, value=tahsin)
        ws.cell(row=siswa_baris, column=tanggal_kolom+3, value=suluk)

        wb.save(FILE_NAME)

        # Reset input
        self.nama_spinner.text = "Pilih Nama"
        self.sabaq_input.text = ""
        self.sabqi_input.text = ""
        self.tahsin_input.text = ""
        self.suluk_input.text = ""

        print("Nilai berhasil disimpan!")

    def hitung_rata_rata_per_murid(self, instance):
        nama = self.nama_spinner.text
        if nama not in siswa_list:
            self.rata_label.text = "Pilih nama siswa!"
            return

        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active

        siswa_baris = None
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == nama:
                siswa_baris = row
                break

        if siswa_baris is None:
            self.rata_label.text = "Data tidak ditemukan!"
            return

        total_sabaq, total_sabqi, total_tahsin, total_suluk = 0, 0, 0, 0
        jumlah_data = 0

        for col in range(2, ws.max_column + 1, 4):
            try:
                sabaq = float(ws.cell(row=siswa_baris, column=col).value or 0)
                sabqi = float(ws.cell(row=siswa_baris, column=col+1).value or 0)
                tahsin = float(ws.cell(row=siswa_baris, column=col+2).value or 0)
                suluk = float(ws.cell(row=siswa_baris, column=col+3).value or 0)

                total_sabaq += sabaq
                total_sabqi += sabqi
                total_tahsin += tahsin
                total_suluk += suluk
                jumlah_data += 1
            except ValueError:
                continue

        if jumlah_data > 0:
            self.rata_label.text = (
                f"Rata-rata {nama}:\n"
                f"Sabaq={total_sabaq/jumlah_data:.2f}, "
                f"Sabqi={total_sabqi/jumlah_data:.2f}, "
                f"Tahsin={total_tahsin/jumlah_data:.2f}, "
                f"Suluk={total_suluk/jumlah_data:.2f}"
            )
        else:
            self.rata_label.text = "Belum ada nilai untuk siswa ini!"

if __name__ == '__main__':
    NilaiTahfizhApp().run()